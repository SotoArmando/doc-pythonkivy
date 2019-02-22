#!/usr/bin/env python
# -*- coding: utf-8 -*-
import kivy
from kivy.config import Config

Config.set('graphics','borderless', 1)
Config.set('graphics','position','custom')
Config.set('graphics','window_state','visible')
Config.set('graphics','resizable',0)
Config.set('graphics','left',500)
Config.set('graphics','top',0)
import threading 
from kivy.garden.mapview import MapView, MapMarker, MarkerMapLayer, MapSource, MapLayer,MapMarkerPopup
from threading import *
from kivy3dgui.layout3d import Layout3D 
from kivy3dgui.layout3d import Node
from kivy.effects.opacityscroll import OpacityScrollEffect
from kivy.effects.scroll import ScrollEffect
from kivy.uix.bubble import Bubble
from kivy.uix.checkbox import CheckBox
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.scatter import Scatter
from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.anchorlayout import AnchorLayout
from kivy.properties import ListProperty, StringProperty
from kivy.adapters.simplelistadapter import SimpleListAdapter
from kivy.adapters.listadapter import ListAdapter
from kivy.uix.listview import ListItemButton, ListView, ListItemLabel,CompositeListItem
from kivy.adapters.models import SelectableDataItem
from kivy.graphics import *
from kivy.core.window import Window
from kivy.utils import get_hex_from_color, get_color_from_hex
from kivy.parser import parse_color
from kivy.graphics.instructions import InstructionGroup
from kivy.uix.accordion import Accordion, AccordionItem
from kivy.base import runTouchApp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.popup import Popup
from kivy.uix.image import Image
from kivy.uix.image import AsyncImage
from kivy.core.window import Window
from kivy.uix.dropdown import DropDown
from kivy.animation import Animation
from kivy.uix.widget import Widget
from kivy.uix.stencilview import StencilView
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.properties import (ObjectProperty, NumericProperty, OptionProperty, BooleanProperty, StringProperty)
from kivy.resources import resource_add_path
from kivy.uix.screenmanager import ScreenManager, Screen,SwapTransition, NoTransition,SlideTransition,FadeTransition,WipeTransition,FallOutTransition,RiseInTransition 
from kivy.uix.behaviors import ButtonBehavior
from kivy.lang import Builder
from kivy.uix.button import Button
from kivy.clock import Clock, mainthread
from kivy.uix.relativelayout import RelativeLayout
from kivy.app import App

import requests
import os   
import os.path
import time
import sys
import urllib2
import json, requests
import math
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from plyer import gps
from threading import Timer



reload(sys)
resource_add_path(os.path.dirname(__file__))

Window.size = (426,950)

Builder.load_string('''
<NavigationDrawer>:
    size_hint: (1,1)
    _side_panel: sidepanel
    _main_panel: mainpanel
    _join_image: joinimage
    side_panel_width: min(0.5*self.width, 0.5*self.width)
    BoxLayout:
        id: sidepanel
        y: root.y
        x: root.x - \
           (1-root._anim_progress)* \
           root.side_panel_init_offset*root.side_panel_width
        height: root.height
        width: root.side_panel_width
        opacity: root.side_panel_opacity + \
                 (1-root.side_panel_opacity)*root._anim_progress
        canvas:
            Color:
                rgba: (0,0,0,.80)
            Rectangle:
                pos: self.pos
                size: self.size
        canvas.after:
            Color:
                rgba: (1,1,1,(1-root._anim_progress)*root.side_panel_darkness)
            Rectangle:
                size: self.size
                pos: self.pos
    BoxLayout:
        id: mainpanel
        x: root.x + \
           root._anim_progress * \
           root.side_panel_width * \
           root.main_panel_final_offset
        y: root.y
        size: root.size
        canvas:
            Color:
                rgba: (1,1,1,1)
            Rectangle:
                pos: self.pos
                size: self.size
        canvas.after:
            Color:
                rgba: (0,0,0,root._anim_progress*root.main_panel_darkness)
            Rectangle:
                size: self.size
                pos: self.pos
    Image:
        id: joinimage
        opacity: min(sidepanel.opacity, 0 if root._anim_progress < 0.00001 \
                 else min(root._anim_progress*40,1))
        source: root._choose_image(root._main_above, root.separator_image)
        mipmap: False
        width: 1
        height: root._side_panel.height
        x: (mainpanel.x - self.width + 1) if root._main_above \
           else (sidepanel.x + sidepanel.width - 1)
        y: root.y
        allow_stretch: True
        keep_ratio: False
''')
class NavigationDrawerException(Exception):
    '''Raised when add_widget or remove_widget called incorrectly on a
    NavigationDrawer.

    '''

class ImageButton(ButtonBehavior, AsyncImage):
    pass
class NavigationDrawer(StencilView):

    # Internal references for side, main and image widgets
    _side_panel = ObjectProperty()
    _main_panel = ObjectProperty()
    _join_image = ObjectProperty()

    side_panel = ObjectProperty(None, allownone=True)
    main_panel = ObjectProperty(None, allownone=True)


    # Appearance properties
    side_panel_width = NumericProperty()

    separator_image = StringProperty('')


    # Touch properties
    touch_accept_width = NumericProperty('14dp')
    _touch = ObjectProperty(None, allownone=True)  # The currently active touch

    # Animation properties
    state = OptionProperty('closed', options=('open', 'closed'))
    anim_time = NumericProperty(0.3)
    min_dist_to_open = NumericProperty(0.7)
    _anim_progress = NumericProperty(0)  # Internal state controlling
                                         # widget positions
    _anim_init_progress = NumericProperty(0)

    # Animation controls
    top_panel = OptionProperty('main', options=['main', 'side'])
    _main_above = BooleanProperty(True)

    side_panel_init_offset = NumericProperty(0.5)

    side_panel_darkness = NumericProperty(0.8)

    side_panel_opacity = NumericProperty(1)

    main_panel_final_offset = NumericProperty(1)

    main_panel_darkness = NumericProperty(0)

    opening_transition = StringProperty('out_cubic')

    closing_transition = StringProperty('in_cubic')

    anim_type = OptionProperty('reveal_from_below',
                               options=['slide_above_anim',
                                        'slide_above_simple',
                                        'fade_in',
                                        'reveal_below_anim',
                                        'reveal_below_simple',
                                        ])

    def __init__(self, **kwargs):
        super(NavigationDrawer, self).__init__(**kwargs)
        Clock.schedule_once(self.on__main_above, 0)

    def on_anim_type(self, *args):
        anim_type = self.anim_type
        if anim_type == 'slide_above_anim':
            self.top_panel = 'side'
            self.side_panel_darkness = 0
            self.side_panel_opacity = 1
            self.main_panel_final_offset = 0.5
            self.main_panel_darkness = 0.5
            self.side_panel_init_offset = 1
        if anim_type == 'slide_above_simple':
            self.top_panel = 'side'
            self.side_panel_darkness = 0
            self.side_panel_opacity = 1
            self.main_panel_final_offset = 0
            self.main_panel_darkness = 0
            self.side_panel_init_offset = 1
        elif anim_type == 'fade_in':
            self.top_panel = 'side'
            self.side_panel_darkness = 0
            self.side_panel_opacity = 0
            self.main_panel_final_offset = 0
            self.main_panel_darkness = 0
            self.side_panel_init_offset = 0.5
        elif anim_type == 'reveal_below_anim':
            self.top_panel = 'main'
            self.side_panel_darkness = 0.8
            self.side_panel_opacity = 1
            self.main_panel_final_offset = 1
            self.main_panel_darkness = 0
            self.side_panel_init_offset = 0.5
        elif anim_type == 'reveal_below_simple':
            self.top_panel = 'main'
            self.side_panel_darkness = 0
            self.side_panel_opacity = 1
            self.main_panel_final_offset = 1
            self.main_panel_darkness = 0
            self.side_panel_init_offset = 0

    def on_top_panel(self, *args):
        if self.top_panel == 'main':
            self._main_above = True
        else:
            self._main_above = False

    def on__main_above(self, *args):
        newval = self._main_above
        main_panel = self._main_panel
        side_panel = self._side_panel
        self.canvas.remove(main_panel.canvas)
        self.canvas.remove(side_panel.canvas)
        if newval:
            self.canvas.insert(0, main_panel.canvas)
            self.canvas.insert(0, side_panel.canvas)
        else:
            self.canvas.insert(0, side_panel.canvas)
            self.canvas.insert(0, main_panel.canvas)

    def toggle_main_above(self, *args):
        if self._main_above:
            self._main_above = False
        else:
            self._main_above = True

    def add_widget(self, widget):
        if len(self.children) == 0:
            super(NavigationDrawer, self).add_widget(widget)
            self._side_panel = widget
        elif len(self.children) == 1:
            super(NavigationDrawer, self).add_widget(widget)
            self._main_panel = widget
        elif len(self.children) == 2:
            super(NavigationDrawer, self).add_widget(widget)
            self._join_image = widget
        elif self.side_panel is None:
            self._side_panel.add_widget(widget)
            self.side_panel = widget
        elif self.main_panel is None:
            self._main_panel.add_widget(widget)
            self.main_panel = widget
        else:
            raise NavigationDrawerException(
                'Can\'t add more than two widgets'
                'directly to NavigationDrawer')

    def remove_widget(self, widget):
        if widget is self.side_panel:
            self._side_panel.remove_widget(widget)
            self.side_panel = None
        elif widget is self.main_panel:
            self._main_panel.remove_widget(widget)
            self.main_panel = None
        else:
            raise NavigationDrawerException(
                'Widget is neither the side or main panel, can\'t remove it.')

    def set_side_panel(self, widget):
        '''Removes any existing side panel widgets, and replaces them with the
        argument `widget`.
        '''
        # Clear existing side panel entries
        if len(self._side_panel.children) > 0:
            for child in self._side_panel.children:
                self._side_panel.remove(child)
        # Set new side panel
        self._side_panel.add_widget(widget)
        self.side_panel = widget

    def set_main_panel(self, widget):
        '''Removes any existing main panel widgets, and replaces them with the
        argument `widget`.
        '''
        # Clear existing side panel entries
        if len(self._main_panel.children) > 0:
            for child in self._main_panel.children:
                self._main_panel.remove(child)
        # Set new side panel
        self._main_panel.add_widget(widget)
        self.main_panel = widget

    def on__anim_progress(self, *args):
        if self._anim_progress > 1:
            self._anim_progress = 1
        elif self._anim_progress < 0:
            self._anim_progress = 0
        if self._anim_progress >= 1:
            self.state = 'open'
        elif self._anim_progress <= 0:
            self.state = 'closed'

    def on_state(self, *args):
        Animation.cancel_all(self)
        if self.state == 'open':
            self._anim_progress = 1
        else:
            self._anim_progress = 0

    def anim_to_state(self, state):
        '''If not already in state `state`, animates smoothly to it, taking
        the time given by self.anim_time. State may be either 'open'
        or 'closed'.

        '''
        if state == 'open':
            anim = Animation(_anim_progress=1,
                             duration=self.anim_time,
                             t=self.closing_transition)
            anim.start(self)
        elif state == 'closed':
            anim = Animation(_anim_progress=0,
                             duration=self.anim_time,
                             t=self.opening_transition)
            anim.start(self)
        else:
            raise NavigationDrawerException(
                'Invalid state received, should be one of `open` or `closed`')

    def toggle_state(self, animate=True):
        '''Toggles from open to closed or vice versa, optionally animating or
        simply jumping.'''
        if self.state == 'open':
            if animate:
                self.anim_to_state('closed')
            else:
                self.state = 'closed'
        elif self.state == 'closed':
            if animate:
                self.anim_to_state('open')
            else:
                self.state = 'open'

    def on_touch_down(self, touch):
        col_self = self.collide_point(*touch.pos)
        col_side = self._side_panel.collide_point(*touch.pos)
        col_main = self._main_panel.collide_point(*touch.pos)

        if self._anim_progress < 0.001:  # i.e. closed
            valid_region = (self.x <=
                            touch.x <=
                            (self.x + self.touch_accept_width))
            if not valid_region:
                self._main_panel.on_touch_down(touch)
                return False
        else:
            if col_side and not self._main_above:
                self._side_panel.on_touch_down(touch)
                return False
            valid_region = (self._main_panel.x <=
                            touch.x <=
                            (self._main_panel.x + self._main_panel.width))
            if not valid_region:
                if self._main_above:
                    if col_main:
                        self._main_panel.on_touch_down(touch)
                    elif col_side:
                        self._side_panel.on_touch_down(touch)
                else:
                    if col_side:
                        self._side_panel.on_touch_down(touch)
                    elif col_main:
                        self._main_panel.on_touch_down(touch)
                return False
        Animation.cancel_all(self)
        self._anim_init_progress = self._anim_progress
        self._touch = touch
        touch.ud['type'] = self.state
        touch.ud['panels_jiggled'] = False  # If user moved panels back
                                            # and forth, don't default
                                            # to close on touch release
        touch.grab(self)
        return True

    def on_touch_move(self, touch):
        if touch is self._touch:
            dx = touch.x - touch.ox
            self._anim_progress = max(0, min(self._anim_init_progress +
                                            (dx / self.side_panel_width), 1))
            if self._anim_progress < 0.975:
                touch.ud['panels_jiggled'] = True
        else:
            super(NavigationDrawer, self).on_touch_move(touch)
            return

    def on_touch_up(self, touch):
        if touch is self._touch:
            self._touch = None
            init_state = touch.ud['type']
            touch.ungrab(self)
            jiggled = touch.ud['panels_jiggled']
            if init_state == 'open' and not jiggled:
                if self._anim_progress >= 0.975:
                        self.anim_to_state('closed')
                else:
                    self._anim_relax()
            else:
                self._anim_relax()
        else:
            super(NavigationDrawer, self).on_touch_up(touch)
            return

    def _anim_relax(self):
        '''Animates to the open or closed position, depending on whether the
        current position is past self.min_dist_to_open.

        '''
        if self._anim_progress > self.min_dist_to_open:
            self.anim_to_state('open')
        else:
            self.anim_to_state('closed')

    def _choose_image(self, *args):
        '''Chooses which image to display as the main/side separator, based on
        _main_above.'''
        if self.separator_image:
            return self.separator_image
        if self._main_above:
            return 'navigationdrawer_gradient_rtol.png'
        else:
            return 'navigationdrawer_gradient_ltor.png'

# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
C1 = "[color=#13C0C7]"
C4 = "[color=#000000]"
C2 = "[color=#404040]"
C3 = "[color=#ff3333]"

def_textsize = (Window.width-50,50)
patch = os.path.dirname(os.path.abspath(__file__))
hud = patch + '/hud/'
icon = patch + '/icons/'
textb = patch + '/textbox/'
database = patch + '/database/'
color = patch + '/hud/color/'
barra = patch + '/hud/barras/'

book1 = load_workbook(filename = database+'CLIENTES.xlsx')
book2 = load_workbook(filename = database+'MUESTRA2.xlsx')
book3 = load_workbook(filename = database+'CXC.xlsx')

sheet1 = book1["Clientes"]
sheet2 = book2["Hoja2"]
sheet3 = book3["Hoja1"]  





class Showcase(FloatLayout):
    pass
def on_size(self, *args):
    self.parent.parent.on_size()
class SideLabel(Label):
    on_size = on_size
class SideButton(Button):
    on_size = on_size

class InterfaceManager(RelativeLayout):
    def __init__(self, **kwargs):
        linea = 0
        book1 = load_workbook(filename = database+'CLIENTES.xlsx')
        book2 = load_workbook(filename = database+'MUESTRA2.xlsx')
        book3 = load_workbook(filename = database+'CXC.xlsx')

        sheet1 = book1["Clientes"]
        sheet2 = book2["Hoja2"]
        sheet3 = book3["Hoja1"]   
        self.Inventarioset = []
        self.Clientesset = []
        Cuentas = []
        self.tablas = [self.Inventarioset,self.Clientesset,Cuentas]
        sheet2set = [sheet2,sheet1,sheet3]
        cont = 0
        for tabla in self.tablas:
            for i in sheet2set[cont]:
                #print "Linea "+str(linea)
                print max
                linea += 1
                columna = 0
                r_linea = []
                for x in i:
                    r_linea.append(str(x.value))
                    #print "Columna "+str(columna) +" = " + str(x.value)
                    columna += 1
                #print r_linea
                tabla.append(r_linea)
            cont+=1
            

            
        
        #print self.Inventarioset
        #print self.Clientesset 
        self.state1 = 1 #Top Menu
        self.state2 = 1 #Bottom Menu
        self.state3 = 1
        self.state4 = 1
        self.state5 = 1 #Sustantivo
        self.state6 = 1 #Adjetivo
        self.state7 = 1 #Nuevo Pedido
        self.state8 = 1 #Pregunta
        self.countstate1 = 0 #Nuevo Pedido
        self.statearray1 = [] #Lista de Compra
        self.statearray2 = [] #Propiedades Locales
        #["ID","Cliente","Fecha","Posicion Actual"] 
        
        super(InterfaceManager, self).__init__(**kwargs)
        self.rangodetexto_cond4 = []
        C1 = "[color=#13C0C7]"
        C4 = "[color=#000000]"
        C2 = "[color=#404040]"
        C3 = "[color=#ff3333]"
        C5 = "[color=#FFD800]"

        Window.clearcolor = (1,1,1,0)
        accesstoken = "pk.eyJ1IjoiYXJtYW5kbzI5IiwiYSI6ImNpd282ZHJ3azAwMWoydHFuZmJudnNzYzEifQ.12vIF51BCThjrut4Q56sGg"
        sourcex = MapSource(url="https://api.mapbox.com/styles/v1/mapbox/streets-v10/tiles/256/{z}/{x}/{y}?access_token=pk.eyJ1IjoibWFjYXR1cyIsImEiOiJjaXlubGRkdXAwMDJ1MzNwZjdwanhkdnd6In0.eYV9WVlCsI44Ku0HSup7Pg",
        cache_key="custom_map1s",tile_size=256,
        image_ext="jpg", attribution="@Armando Jose Soto Melo")
        self.mv = MapView(zoom = 15, lat = 18.454651 , lon = -69.971119, map_source = sourcex)
        self.mv_container = Layout3D(size_hint = (1,1),)
        m1 = MapMarker(lat = 18.454651 , lon = -69.971119)  # Lille
        self.mv.add_marker(m1)
        
        self.adjetivosroot = GridLayout(cols = 1, size_hint_y = None, height = Window.height, size_hint_x = None , width = 1000)
        self.adjetivosroot_root = ScrollView(size_hint_y = None, height = Window.height, pos = (Window.width,0), do_scroll_x = True)
        self.adjetivosroot_root.add_widget(self.adjetivosroot)
        self.adjetivosroot.bind(minimum_height = self.adjetivosroot.setter('height'))        
        self.adjetivosroot.bind(minimum_width = self.adjetivosroot.setter('width'))        
        self.adjetivosroot.add_widget(Button(background_normal = color + "16.png",size_hint_y = None, height = 50))
        self.adjetivosroot.add_widget(Button(markup = True,text = C4+"Estado de cuentas",background_normal = color + "16.png",size_hint_y = None, height = 50))
        self.adjetivos_columns = GridLayout(cols = 7, size_hint_y = None, size_hint_x = None , width = 1000, height = 10*50)
        self.adjetivos_columns1 = GridLayout(cols = 6, size_hint_y = None,height =150, size_hint_x = None , width = 1000)
        self.adjetivosroot.add_widget(self.adjetivos_columns)
        self.adjetivosroot.add_widget(self.adjetivos_columns1)
        Button11 = Button(markup = True,text = C4+"Volver",background_normal = color + "1.png",size_hint_y = None, height = 50)
        self.adjetivosroot.add_widget(Button11)
        Button11.bind(on_release = lambda x :self.adjetivo())

        for i in range(10):
            if i == 0:
                set7_un = ["Documento","Emision","Vencto.","Dias","Balance Original","Creditos o Pagos","Balance Actual"]
                set7 = []
                for i in set7_un:
                    set7.append([3.02,i])
                self.MasterWidget(self.adjetivos_columns,set7)
                
            set6 = ["12671","20/11/2015","20/12/2015","376","19,250.00","6,040.00","13,210.00"]
            self.MasterWidget(self.adjetivos_columns,set6)
        
        
        set8_un = ["","0 30","31 - 60","61 - 90","91 - 120","121 - ..."]
        set8 = []
        for i in set8_un:
            set8.append([3.02,i])
            
        set9 = ["Balance","","","","",""]
        set10 = ["Porciento","","","","",""]
        self.MasterWidget(self.adjetivos_columns1,set8)
        self.MasterWidget(self.adjetivos_columns1,set9)
        self.MasterWidget(self.adjetivos_columns1,set10)
        
        self.sustantivosroot = GridLayout(cols = 1, size_hint_y = None, height = Window.height)
        self.sustantivosroot_root = ScrollView(cols = 1, size_hint_y = None, height = Window.height, pos = (Window.width,0))
        self.sustantivosroot_root.add_widget(self.sustantivosroot)
        self.sustantivosroot.bind(minimum_height = self.sustantivosroot.setter('height'))        
        self.sustantivosroot.add_widget(Button(background_normal = color + "16.png",size_hint_y = None, height = 50))
        Button10 = Button(background_normal = color + "1.png",size_hint_y = None, height = 50)
        self.sustantivosroot.add_widget(Button10)
        
        for i in range(10):
            set4 = [[6.01,"TRANSPORTE ANDROMEDA SA",""]]
            self.MasterWidget(self.sustantivosroot,set4)
            set5 = [[3.01,C5+"Balance Original[/color]800,000.00 RD$"],[3.01,C5+"Balance Actual[/color]125,000.00 RD$"],[3.02,C5+"Cantidad de Documentos[/color] 15"],[3.02,C5+"No.[/color] 00024"]
            ]
            dim = GridLayout(cols = 2, size_hint_y = None, height = 150)
            self.MasterWidget(dim,set5)
            self.sustantivosroot.add_widget(dim)
            self.sustantivosroot.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
        Button10.bind(on_release = lambda x: self.sustantivo())   
        
        
        self.menu_posy = 70
        self.menu = GridLayout(cols = 1, size_hint = (None,None), size = (Window.width,450) , pos = (0,Window.height-self.menu_posy), spacing = 0)
        self.screenmanager = ScreenManager(size_hint = (1,1), size = (Window.width,Window.height) , pos = (0,-Window.height+70), spacing = 0)
        self.screen1 = Screen(name = "a")
        self.screen2 = Screen(name = "b")
        self.screen3 = Screen(name = "c")
        self.screenmanager.add_widget(self.screen1)
        self.screenmanager.add_widget(self.screen2)
        self.screenmanager.add_widget(self.screen3)
        
        self.screen3root = GridLayout(cols = 1, size_hint_y = None, height = Window.height)
        self.screen3root_root = ScrollView(cols = 1, size_hint_y = None, height = Window.height)
        self.screen3root_root.add_widget(self.screen3root)
        self.screen3root.bind(minimum_height = self.screen3root.setter('height'))        
        self.screen3.add_widget(self.screen3root_root)
        self.screen3root.add_widget(Button(background_normal = color + "16.png",size_hint_y = None, height = 50))
        self.screen3root_lista = GridLayout(cols = 1, size_hint_y = None)
        self.screen3root.add_widget(self.screen3root_lista)

            
        self.screen3root.add_widget(Button(background_normal = color + "16.png",size_hint_y = None, height = 50))
        set3 = [[1,"Documento No.",True],[1,"Posicion del Mapa",True],[1,"Cliente",True],[1,"Emision",True],[1,"Vencto",True],[1,"Dias"],[1,"Balance Original",True],[1,"Creditos o Pagos"],[1,"Balance Actual",True]]
        
        self.MasterWidget(self.screen3root,set3)

        deshabilitados = [17,15,13,11,9,5,1]
        #ID,Latlon,Cliente,FechaC,FechaV,BalanceO,BalanceA
        correspondientes = ["1","18.99,17.88","Armando Jose Soto Melo","05/02/2017","05/03/2017","800,000.00 RD$","0.00 RD$"]
        c = 0
        for i in deshabilitados:
            self.screen3root.children[i].children[0].text = correspondientes[c]
            c += 1

        Button9 = Button(markup = True,text = C4+"CONFIRMAR",background_normal = color + "6.png",size_hint_y = None, height = 50)
        
        self.screen3root.add_widget(Button9)
        Button9.bind(on_release = lambda x: self.Next_Screen())
        self.dim3 = GridLayout(cols = 1, size_hint_y = None, height = Window.height)
        self.dim3_root = ScrollView(cols = 1, size_hint_y = None, height = Window.height)
        self.dim3_root.add_widget(self.dim3)
        self.dim3.bind(minimum_height = self.dim3.setter('height'))        
        self.dim3.add_widget(Button(background_normal = color + "16.png",size_hint_y = None, height = 50))
        
        
        
        for i in range(len(self.Inventarioset)):
            self.dim4 = GridLayout(cols = 2, size_hint_y = None, height = 300)
            self.dim4a = GridLayout(cols = 1 , size_hint_y = None , height = 300)
            self.dim4b = GridLayout(cols = 1 , size_hint_y = None , height = 300)
            self.dim4.add_widget(self.dim4a)
            self.dim4.add_widget(self.dim4b)
            textset1 = ["CODIGO","DESCRIPCION","MARCA","CLASIFICACION"]
            textset2 = ["TOTAL EXISTENCIA","PRECIO #1","PRECIO #2","PRECIO #3"]
            for x in range(4):
                Button2 = Button(text = C1+textset1[x]+"[/color]"+" :"+str(self.Inventarioset[i][x]),background_normal = color+"3.png", text_size = (Window.width/2 -50, 75), valign = "middle", halign = "left",markup = True)
                Button3 = Button(text = C1+textset2[x]+"[/color]"+" :"+str(self.Inventarioset[i][4+x]),background_normal = color+"3.png", text_size = (Window.width/2 -50, 75), valign = "middle", halign = "left",markup = True)
                self.dim4a.add_widget(Button2)
                self.dim4b.add_widget(Button3)


            self.Button4 = Button(markup = True,text = C4+"Descripcion :"+str(self.Inventarioset[i][1]),size_hint_y = None,height = 50,background_normal = color+"16.png", text_size = (Window.width -100, 50), valign = "middle", halign = "left")
            dim6 = GridLayout(rows = 1, size_hint_y = None, height = 50)
            self.Button7 = Button(background_normal = color + "6.png", text = str(i), font_size = 0, size_hint_x = None, width = 50)
            self.Button7.bind(on_release =lambda x: self.agregarobjeto(x.text))
            dim6.add_widget(self.Button4)
            dim6.add_widget(self.Button7)
            self.dim3.add_widget(dim6)
            self.dim3.add_widget(self.dim4)
            self.dim3.add_widget(Image(source = color + "2.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
            
        self.screen2.add_widget(self.dim3_root)
        self.nuevo_cliente = GridLayout(cols = 1, size_hint = (None,None), size = (Window.width,Window.height), spacing = 0)
        menuset = ["Menu","Nuevo Pedido","Tirar Ubicacion","","","","Menu"]
        
        menusetx1 = [["buscador","Menu Contextual","10"],[3.04,"Pedido","Tocar para crear."],[3.04,"Geolocalizar Dispositivo",""],[3.04,"CxC","Cuentas por Cobrar"],[3.04,"Historial",""]]
        self.MasterWidget(self.menu,menusetx1) 
            
        for i in range(1):
            if i == 0:
                pass
 
                #self.menu.add_widget(Image(source = color + "16c.png", size_hint_y= None , height =1, allow_stretch = True , keep_ratio = False))
            #self.menu.add_widget(Button(size_hint = (None,None), size = (50,50), pos = (0,0)))
            Button1 =Button(font_size = 19,size_hint_y = None, height = 50,text_size = (Window.width-50,50),valign = "middle",text =C4+ menuset[i], background_normal = color+"16b.png", markup = True)
            
            self.menu.add_widget(Button1)
            
            Button1.bind(on_release = lambda x: self.Animate(self.menu,(0,Window.height-self.menu_posy),(0,Window.height-450)))
            Button1.background_down =  color+"16c.png"
            
            if i == 2:
                Button1.bind(on_release = lambda x: self.ubicar())
            elif i == 3:
                Button1.bind(on_release = lambda x: self.sustantivo())
            elif i == 4:
                Button1.bind(on_release = lambda x: self.Nuevopedido())
                
        menusetx = [[7,'']]
        self.MasterWidget(self.menu,menusetx)    
        menusetx2 = [[7.01,'']]
        self.MasterWidget(self.nuevo_cliente,menusetx2)       
        menuset1 = ["Nuevo Cliente",[1,"Cliente"],[1,"Identificacion"],[2,"RNC","CEDULA"],[1,"Contacto"],[1,"Telefono1"],[1,"Telefono2"],[1,"Direccion1"],[1,"Sector"]
        ,[1,"Ciudad"],[1,"Direccion 2"],[1,"Sector"],[1,"Ciudad"]]
        for i in range(13):
            try:
                if menuset1[i][0] == 1:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,text_size = (100,60), halign = "left", valign = "middle",size_hint_y = None,size_hint_x = None, width = 150, height = 50,markup = True,text = C4+menuset1[i][1]+":", background_normal = color+"16.png"))
                    dim1.add_widget(TextInput(padding = [25,15,0,0],size_hint_y = None,size_hint_x = 1, width = 100, height = 50,markup = True,text = "", background_normal = color+"16b.png", background_active = color+"16b.png"))
                    self.nuevo_cliente.add_widget(dim1)
                    self.nuevo_cliente.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset1[i][0] == 2:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(ToggleButton(text_size = (Window.width/2 -50,60), halign = "center", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = "[b]"+C4+menuset1[i][1]+"", background_normal = color+"16.png",background_down = color+"1.png",group = "a"))
                    dim1.add_widget(ToggleButton(text_size = (Window.width/2 -50,60), halign = "center", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = "[b]"+C4+menuset1[i][2]+"", background_normal = color+"16.png",background_down = color+"1.png",group = "a"))
                    self.nuevo_cliente.add_widget(dim1)
                    self.nuevo_cliente.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))

                else:
                    print i
                    Button1 =Button(font_size = 19,size_hint_y = None, height = 50,markup = True,text =C4+ menuset1[i], background_normal = color+"16.png", background_down = color+"16.png")
                    self.nuevo_cliente.add_widget(Button1)
                    self.nuevo_cliente.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
            except:
                print i
                Button1 =Button(font_size = 19,size_hint_y = None, height = 50,markup = True,text =C4+ menuset1[i], background_normal = color+"16.png", background_down = color+"16.png")
                self.nuevo_cliente.add_widget(Button1)
                self.nuevo_cliente.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
            
            if i == 0:
                Button1.bind(on_release = lambda x: self.Animate1(self.screenmanager,(0,-Window.height+70),(0,0)))
            if i == 12:
                dim2 = GridLayout(rows = 1, height = 50, size_hint_y = None )
                buttonset = ["Aprobar","Cancelar"]
                colorset = ["16","16"]
                for x in range(2):
                    button2 =Button(font_size = 19,text_size = (Window.width/2 -50,60), halign = "center", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = C4+buttonset[x]+"", background_normal = color+colorset[x]+".png")
                    dim2.add_widget(button2)
                    if x == 1:
                        button2.bind(on_release = lambda x: self.Animate1(self.screenmanager,(0,-Window.height+70),(0,0)))
                    elif x == 0:
                        button2.bind(on_release = lambda x: self.agregarcliente())
                        
                    
                    
                self.nuevo_cliente.add_widget(dim2)
        self.nuevo_cliente.add_widget(Button(size_hint= (1,None), height = 20,background_normal = barra + "barra1.png"))


         

        self.busquedagrid = GridLayout(cols = 1,size_hint = (None,None),width = Window.width, height = 550, pos = (0,Window.height+10))
      
        
        Button5 = Button(background_normal = color + "5.png",text = C4+"Panel de Busqueda",background_down = color + "5.png",size_hint_y = None, height = 50,markup = True)
        Button5.bind(on_release = lambda x:self.Animate2(self.busquedagrid,(0,Window.height-40),(0,Window.height-560)))
        set = ["Panel de Busqueda",[4,"Codigo",0],[4,"Descripcion",1],[4,"Marca",2],[4,"Clasificacion",3],[4,"Total Existencia",4],
        [4,"Precio 1",5],
        [4,"Precio 2",6],
        [4,"Precio 3",7],"Realizar Busqueda"]
        self.MasterWidget(self.busquedagrid,set)
       
        
        #Cliente ,Identificacion, Contacto, Telefono1, Telefono2, Direccion1, Sector, Ciudad, Direccion2, Sector,Ciudad
        #CODIGO	CLIENTE	  RNC/CEDULA	CONTACTO	TELEFONO1	TELEFONO2	DIRECCION1	SECTOR	CIUDAD	DIRECCION2	SECTOR	CIUDAD



        
        
        
        
        self.Login = GridLayout(cols = 1)
        self.Login.add_widget(Button(background_normal = color + "16.png"))
        self.Login.add_widget(Button(text_size = (Window.width-50,32.5),background_normal = color + "16.png", size_hint_y = None , height = 75, font_size = 24 ,markup = True, text = C4+"Inicio"))
        self.Login.add_widget(Button(text_size = (Window.width-50,32.5),background_normal = color + "16b.png", size_hint_y = None , height = 75, font_size = 24 ,markup = True, text = C4+""))
        set1 = ["Usuario","Clave"] 
        for i in range(2):
            dim5 = GridLayout(rows = 1, size_hint_y = None, height = 50)
            dim5.add_widget(Button(text_size = (100,50), valign = "middle",markup = True,text = C4+set1[i]+" :",background_normal = color + "16b.png",background_down = color + "16b.png",size_hint_x = None, width = 150))
            dim5.add_widget(TextInput(padding = [25,15,0,0],text = set1[i],background_normal = color + "16c.png",background_active = color + "16.png"))
            self.Login.add_widget(dim5)
        self.Login.add_widget(Button(text_size = (Window.width-50,32.5),background_normal = color + "16b.png", size_hint_y = None , height = 75, font_size = 24 ,markup = True, text = C4+""))
        set = [[5,"text"]]
        self.MasterWidget(self.Login,set)
        self.Login.add_widget(Button(background_normal = color + "16.png"))
        dim7 = GridLayout(rows = 1,size_hint_y= None , height = 50)
        Button8 = Button(size_hint_x = None, width = 50, background_normal = color +"10.png")
        Button8.bind(on_release = lambda x: self.Next_Screen())
        dim7.add_widget(Button5)
        dim7.add_widget(Button8)
        self.busquedagrid.add_widget(dim7)
        self.add_widget(self.mv)
        self.add_widget(self.menu)
        self.locationtext = Button(border = [0,0,0,0],markup = True,background_normal = barra + "barrarad1.png",background_down = barra + "barrarad1.png",size_hint = (None,None), size = (Window.width,50),text_size = (Window.width,50),halign = "center",valign = "middle", pos = (0,350))
        self.add_widget(self.locationtext)
        self.hstate_root = RelativeLayout(size_hint = (1,None), height = 350, pos = (0,0))
        self.hstate = ScrollView(size_hint = (1,None), height = 350, pos = (0,0))
        self.hstate_child = GridLayout(cols = 1, size_hint = (1,None))
        self.hstate.add_widget(self.hstate_child)
        self.hstate_child.bind(minimum_height = self.hstate_child.setter('height'))
        self.hstate_root.add_widget(Button(background_normal = color + "16.png"))
        self.hstate_root.add_widget(self.hstate)
        self.add_widget(self.hstate_root)
        menusetx3 = [[3.05,"Pedido","Direccion"],[3.05,"Pedido","Direccion"],[3.05,"Pedido","Direccion"],[3.05,"Pedido","Direccion"],[3.05,"Pedido","Direccion"],[3.05,"Pedido","Direccion"]]
        self.MasterWidget(self.hstate_child,menusetx3) 
        self.hstate_child.add_widget(Image(size_hint_y = None,height = 50,source = color + "16.png", keep_ratio = False, allow_stretch = True))
        
        
        self.add_widget(self.screenmanager)
        self.add_widget(self.busquedagrid)
        self.add_widget(self.Login)
        self.screen1.add_widget(self.nuevo_cliente)
        self.nuevopedido = GridLayout(cols = 1, size_hint_y = None , height  = 100, pos = (Window.width,Window.height/2))
        textset3_un = ["Usar un cliente existente.","Crear un nuevo cliente.","Cancelar"]
        textset3 = []
        for i in textset3_un:
            if i == "Cancelar":
                textset3.append([6,i,0.03])
            else:
                textset3.append([6,i])
        self.MasterWidget(self.nuevopedido,textset3)
        self.add_widget(self.nuevopedido)
        
        self.pregunta = GridLayout(cols = 1, size_hint_y = None , height  = 100, pos = (Window.width,Window.height/2))
        textset4_un = ["Desea crear un pedido con el usuario nuevo?","Si","No"]
        textset4 = []
        for i in textset4_un:
            if i == "No":
                textset4.append([6,i,0.02])
            elif i == "Si":
                textset4.append([6,i,0.04])
            else:
                textset4.append([6,i])
                
        print textset4
        self.MasterWidget(self.pregunta,textset4)
        self.add_widget(self.pregunta)
        
        gps_location = StringProperty()
        gps_status = StringProperty('press to get GPS location updates')

        self.Notificar("Se ha agregado a la lista.")
        try: 
            gps.configure(on_location=self.on_location,on_status=self.on_status)
            
        except NotImplementedError:
            import traceback
            #traceback.print_exc()
            self.gps_status = 'GPS is not implemented for your platform'
            self.locationtext.text = C4 + self.gps_status
            print 'GPS is not implemented for your platform'
    def agregarcliente(self):
        
        childset =[24,22,18,16,14,12,10,8,6,4,2]
        id = self.countstate1 + len(self.Clientesset) 
        row = []
        row.append(id)
        for i in childset:
            row.append(self.nuevo_cliente.children[i+1].children[0].text)
        #self.Animate1(self.screenmanager,(0,-Window.height+50),(0,0))
        
        try:    
            sheet1.append(row)
            book1.save(database+'CLIENTES.xlsx')   
        except:
            import traceback
            traceback.print_exc()
        self.countstate1 += 1
        self.Pregunda1()
    def agregarobjeto(self,index):
        self.statearray1.append(int(index))
        self.Notificar("Se ha agregado al pedido del cliente.")
        print self.statearray1

    def adjetivo(self):
        try:
            self.add_widget(self.adjetivosroot_root)
        except:
            pass
        
        if self.state6 == 1:
            a = Animation(pos = (0,0), d = .5,t='in_out_sine')
            a.start(self.adjetivosroot_root)
        else:
            a = Animation(pos = (Window.width,0), d = .5,t='in_out_sine')
            a.start(self.adjetivosroot_root)

        self.state6 *= -1
    def sustantivo(self):
        try:
            self.add_widget(self.sustantivosroot_root)
        except:
            pass
        
        if self.state5 == 1:
            a = Animation(pos = (0,0), d = .5,t='in_out_sine')
            a.start(self.sustantivosroot_root)
        else:
            a = Animation(pos = (Window.width,0), d = .5,t='in_out_sine')
            a.start(self.sustantivosroot_root)

        self.state5 *= -1
        

    def MasterWidget(self,widget,texset):
        menuset2 = texset
        for i in range(len(menuset2)):

            try:
                if menuset2[i][0] == 1:
                    try:
                        guarda = menuset2[i][2]
                    except:
                        guarda = False
                        
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,text_size = (100,60), halign = "left", valign = "middle",size_hint_y = None,size_hint_x = None, width = 150, height = 50,markup = True,text = C4+menuset2[i][1]+":", background_normal = color+"16.png"))
                    dim1.add_widget(TextInput(padding = [25,15,0,0],size_hint_y = None,size_hint_x = 1, width = 100, height = 50,markup = True,text = "", background_normal = color+"16b.png", background_active = color+"16b.png", disabled = guarda))
                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 2:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(ToggleButton(text_size = (Window.width/2 -50,60), halign = "center", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = "[b]"+C4+menuset2[i][1]+"", background_normal = color+"16.png",background_down = color+"1.png",group = "a"))
                    dim1.add_widget(ToggleButton(text_size = (Window.width/2 -50,60), halign = "center", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = "[b]"+C4+menuset2[i][2]+"", background_normal = color+"16.png",background_down = color+"1.png",group = "a"))
                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 3:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,markup = True,text = menuset2[i][1],background_normal = color +"3.png", slide_hint_y = None, height = 50))
                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                    
                elif menuset2[i][0] == "buscador":
                    dim1 = RelativeLayout(size_hint_y = None, height = 75)
                    dim1.add_widget(Image(source = color+"16.png", allow_stretch = True , keep_ratio = False))
                    
                    dim1b = GridLayout(rows = 1 , size_hint_y = None , height = 75)
                    dim1b.add_widget(TextInput(font_size = 20, text = "Clientes Registrados",padding = [25,25,0,0],background_normal = color+"None.png"))
                    dim1b.add_widget(Image(source = icon+"file-2.png",size = (75,75) ,size_hint = (None,None),pos = (75/2 -20,75/2 -20)))
                    
                    dim1.add_widget(dim1b)
                    

                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))

                elif menuset2[i][0] == 3.04:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 74, spacing = 0)
                    dim1a = RelativeLayout(source = icon+"folder-11.png",size_hint_x = None , width = 75 )
                    dim1a.add_widget(Image(source = color+"16.png",size_hint_x = None , width = 75 ))
                    dim1a.add_widget(Image(source = icon+"folder-11.png",size = (40,40) ,size_hint = (None,None),pos = (75/2 -20,75/2 -20)))
                    dim1_dim = GridLayout(cols = 1,size_hint_y = None , height =75)
                    dim1.add_widget(dim1a)
                    dim1.add_widget(dim1_dim)
                    dim1_dim.add_widget(Button(text_size = (Window.width-75,75),valign = "middle",font_size = 20,background_normal = "16.png", text = C4+menuset2[i][1]+'\n'+"[size=15]"+menuset2[i][2],markup = True))
                    
                    #dim1_dim.add_widget(Button(text_size = (Window.width-75,75/2),valign = "top", font_size = 17,background_normal = "16.png", text = C4+"texto",markup = True))
                    

                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 3.05:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 74, spacing = 0)
                    dim1a = RelativeLayout(size_hint_x = None , width = 75 )
                    dim1a.add_widget(Image(source = color+"16.png",size_hint_x = None , width = 75 ))
                    dim1a.add_widget(Image(source = icon+"briefcase.png",size = (40,40) ,size_hint = (None,None),pos = (75/2 -20,75/2 -20)))
                    dim1_dim = GridLayout(cols = 1,size_hint_y = None , height =75)
                    dim1.add_widget(dim1a)
                    dim1.add_widget(dim1_dim)
                    dim1_dim.add_widget(Button(text_size = (Window.width-75,75),valign = "middle",font_size = 20,background_normal = "16.png", text = C4+menuset2[i][1]+'\n'+"[size=15]"+menuset2[i][2],markup = True))
                    
                    #dim1_dim.add_widget(Button(text_size = (Window.width-75,75/2),valign = "top", font_size = 17,background_normal = "16.png", text = C4+"texto",markup = True))
                    

                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 3.03:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,markup = True,text =C4+ menuset2[i][1],background_normal = color +menuset2[i][2]+".png", slide_hint_y = None, height = 50))
                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 3.01:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 100)
                    dim1.add_widget(Button(font_size = 19,markup = True, text_size = (Window.width/2 -50,100) ,valign = "middle", halign = "left",text = menuset2[i][1],background_normal = color +"3.png", slide_hint_y = None, height = 100))
                    widget.add_widget(dim1)
                    #widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 3.02:
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,markup = True,text = menuset2[i][1],background_normal = color +"3.png", slide_hint_y = None, height = 50))
                    widget.add_widget(dim1)
                    #widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == "b":
                    
                    bar = Image(source = color+menuset2[i][2]+".png", size_hint_y = None, height = 1 )
                    widget.add_widget(bar)
                    #widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 4:

                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    dim1.add_widget(Button(font_size = 19,text_size = (100,60), halign = "left", valign = "middle",size_hint_y = None,size_hint_x = None, width = 150, height = 50,markup = True,text = C4+menuset2[i][1]+":", background_normal = color+"16.png"))
                    
                    TextInput1 =TextInput(padding = [25,15,0,0],size_hint_y = None,size_hint_x = 1, width = 100, height = 50,markup = True,text = "", background_normal = color+"16b.png", background_active = color+"16b.png")
                    dim1.add_widget(TextInput1)
                    
                    ToggleButton1 = ToggleButton(font_size = 0,text_size = (50,50),group = "x", halign = "left", valign = "middle",size_hint_y = None,size_hint_x = None, width = 50, height = 50,markup = True,text = str(menuset2[i][2]), background_normal = color+"16.png", background_down = color+"7.png")
                    self.rangodetexto_cond4.append(TextInput1)
                    ToggleButton1.bind(on_press = lambda x :self.in_state(x.text))
                    dim1.add_widget(ToggleButton1)
                    
                    widget.add_widget(dim1)
                    widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                elif menuset2[i][0] == 5:
                        
                    dim1 = GridLayout(rows = 1,size_hint_y = None, height = 50)
                    Button6 = Button(font_size = 19,text_size = (100,60), halign = "left", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = C4+"Ingresar", background_normal = color+"16b.png", background_down = color+"16b.png")
                    dim1.add_widget(Button6)
                    Button6.bind(on_release = lambda x:self.Main_Removewidget())
                    dim1.add_widget(Button(font_size = 19,text_size = (100,60), halign = "left", valign = "middle",size_hint_y = None,size_hint_x = 1, width = 150, height = 50,markup = True,text = C4+"Salir", background_normal = color+"1.png", background_down = color+"1.png"))
                    widget.add_widget(dim1)
                elif menuset2[i][0] == 7: 
                    dim1_image = Image(size_hint_y = None , height = 20 , source = barra+"barra1.png", keep_ratio = False, allow_stretch = True)
                    widget.add_widget(dim1_image)
                elif menuset2[i][0] == 7.01: 
                    dim1_image = Image(size_hint_y = None , height = 20 , source = barra+"barra2.png", keep_ratio = False, allow_stretch = True)
                    widget.add_widget(dim1_image)
                
                elif menuset2[i][0] == 6: 
                    
                    dim1 = GridLayout(cols =1,height=50,size_hint_y=None)
                    button6 =Button(font_size = 19,text = C4+menuset2[i][1],text_size = def_textsize, valign = "middle",markup = True,background_normal = color + "16.png")
                    dim1.add_widget(button6)
                    widget.add_widget(dim1)
                    
                    try:    
                        if menuset2[i][2] == 0.01:
                            button6.bind(on_release = lambda x: self.Nuevopedido())

                        elif menuset2[i][2] == 0.02:
                            print ("asdasdasd")
                            button6.bind(on_release = lambda x: self.Pregunda1())
                        elif menuset2[i][2] == 0.03:
                            print ("asdasdasd")
                            button6.bind(on_release = lambda x: self.Nuevopedido())
                        elif menuset2[i][2] == 0.04:
                            print ("asdasdasd")
                            button6.bind(on_release = lambda x: self.PedidoETC1())
                        
                        
                    
     
                    except:
                        pass

                  
                elif menuset2[i][0] == 6.01: 
                    dim1 = GridLayout(rows =1,height=50,size_hint_y=None)
                    dim1.add_widget(Button(font_size = 19,text = C4+menuset2[i][1],text_size = (Window.width/2 -50,50), valign = "middle",markup = True,background_normal = color + "16.png"))
                    adjev = Button(font_size = 19,text = C4+menuset2[i][2],text_size =(100,50) ,size_hint_x = None, width = 100,halign = "center", valign = "middle",markup = True,background_normal = color + "6.png")
                    dim1.add_widget(adjev)
                    widget.add_widget(dim1)
                    adjev.bind(on_release = lambda x: self.adjetivo())
                
                    

                
                else:
                    print i
                    Button1 =Button(size_hint_y = None, height = 50,markup = True,text =C4+ menuset2[i], background_normal = color+"16.png", background_down = color+"16.png")
                    widget.add_widget(Button1)
                    #widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))
                    if menuset2[i] == "Realizar Busqueda":
                        Button1.bind(on_release = lambda x:self.filtrar_tabla(0,int(self.state4),self.rangodetexto_cond4))
                        
                    
                        
            except Exception as e:
                import traceback
                traceback.print_exc()
                #print str(e)
                #print i
                Button1 =Button(size_hint_y = None, height = 50,markup = True,text =C4+ menuset2[i], background_normal = color+"16.png", background_down = color+"16.png")
                widget.add_widget(Button1)
                #widget.add_widget(Image(source = color + "16c.png", size_hint_y = None, height = 1, allow_stretch = True , keep_ratio = False))

    def PedidoETC1(self):
        self.Next_Screen()
        self.Pregunda1()
    def ubicar(self):
        a = Animation(d=.5,t='in_out_sine')
        a.start(self.mv)
        print ("animando")
        self.mv.center_on(18.454651,-69.971119)
        
        
    def Notificar(self,texto):
        Notificarbtn = Button(background_normal = color + "T50.png",markup = True,text_size = (Window.width-50,50),valign = "middle",text = texto ,pos = (0,-50),size_hint = (None,None), size = (Window.width,50))
        self.add_widget(Notificarbtn)
        a1 = Animation(pos = (0,0),t='in_out_sine', d = .5)
        a05 = Animation(pos = (0,0),t='in_out_sine', d = 2)
        a2 = Animation(pos = (0,-50),t='in_out_sine', d = .5)
        anim = a1+a05+a2
        anim.start(Notificarbtn)
    def in_state(self,xa):
        self.state4 = xa
        print self.state4
    def Nuevopedido(self):
        
        if self.state7 == 1:
            a = Animation(pos = (0,Window.height/2),d = .5, t= 'in_out_sine')
            a.start(self.nuevopedido)
        else:
            a = Animation(pos = (Window.width,Window.height/2),d = .5, t= 'in_out_sine')
            a.start(self.nuevopedido)
        self.state7 *= -1
    def Pregunda1(self):
        
        if self.state8 == 1:
            a = Animation(pos = (0,Window.height/2),d = .5, t= 'in_out_sine')
            a.start(self.pregunta)
        else:
            a = Animation(pos = (Window.width,Window.height/2),d = .5, t= 'in_out_sine')
            a.start(self.pregunta)
        self.state8 *= -1
    
    def filtrar_tabla(self,tablaindex,campoindex,textoindex):
    
        #print self.rangodetexto_cond4
        self.dim3.clear_widgets()
        tablaseleccionada = self.tablas[tablaindex]
        camposeleccionado = campoindex
        
        datobuscado = textoindex[campoindex].text

        print datobuscado
        for i in tablaseleccionada: 
            try:
                i[camposeleccionado].index(datobuscado)
                print i[camposeleccionado]
            except:
                pass
            
    def Animate(self,button,pos1,pos2):
        self.state1 *= -1
        print self.state1
        if self.state1 == -1:
            button.pos = pos1
            a = Animation(pos = pos2, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text  = C4+"Cerrar"
        elif self.state1 == 1:
            button.pos = pos2
            a = Animation(pos = pos1, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text = C4+"Menu"
    def Animate1(self,button,pos1,pos2):
        self.state2 *= -1
        print self.state2
        if self.state2 == -1:
            button.pos = pos1
            a = Animation(pos = pos2, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text  = "Cerrar"
        elif self.state2 == 1:
            button.pos = pos2
            a = Animation(pos = pos1, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text = "Menu"
    def Animate2(self,button,pos1,pos2):
    
        self.state3 *= -1
        print self.state3
        if self.state3 == -1:
            button.pos = pos1
            a = Animation(pos = pos2, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text  = C4+"Cerrar"
        elif self.state3 == 1:
            button.pos = pos2
            a = Animation(pos = pos1, d = .5, t = "out_cubic")
            a.start(button)
            if pos2 == (0,0):
                pass
            else:
                button.children[0].text = C4+"Panel de Busqueda"
    
    def Animatesize(self,button,p,size2):
        
        print button.height
        button.height = 150
        print button.height
        
        
        a = Animation(height = size2, d = .5, t = "out_cubic")
        #x = Animation(height = size2, d = 5, t = "out_cubic")
        b = Animation(height = size1, d = .5, t = "out_cubic")
        c = a  + b
        c.start(button)

    def Main_Removewidget(self):
        self.remove_widget(self.Login)
        
    def Next_Screen(self):
        print self.screenmanager.next()
        self.screenmanager.current = self.screenmanager.next()
        if self.screenmanager.current == "b":
            try:
                self.add_widget(self.busquedagrid)
            except: pass
            print ("aaa")
            self.busquedagrid.pos = (0,Window.height-40)
        elif self.screenmanager.current == "c": 
            print("aaa2")
            self.remove_widget(self.busquedagrid)
            self.add_widget(self.busquedagrid)
            self.busquedagrid.pos = (0,Window.height-40)
            c = 0
            self.screen3root_lista.clear_widgets()
            for i in self.statearray1:
                c += 1
                set2 = [[6,"Descripcion: "+self.Inventarioset[i][1]],[6,"Precio: "+self.Inventarioset[i][5]]
                ]
                self.MasterWidget(self.screen3root_lista,set2)
                self.screen3root_lista.add_widget(Image(source = color + "16b.png", allow_stretch = True, keep_ratio = False, size_hint_y = None, height =1))
            self.screen3root_lista.height = c*100
        elif self.screenmanager.current == "a": 
            self.remove_widget(self.busquedagrid)
            self.Notificar("Pedido completado con exito.")
        else:   
            self.busquedagrid.pos = (0,Window.height)
    def start(self, minTime, minDistance):
        gps.start()

    def stop(self):
        gps.stop()

    @mainthread
    def on_location(self, **kwargs):
        self.gps_location = '\n'.join([
            '{}={}'.format(k, v) for k, v in kwargs.items()])
            
        a = '\n'.join([
            '{}={}'.format(k, v) for k, v in kwargs.items()])
        for k, v in kwargs.items():
            if (k == "lon") or (k == "lat"):
                print str(k)+"="+str(v)
        self.locationtext.text = C4 + self.gps_location
    @mainthread
    def on_status(self, stype, status):
        self.gps_status = 'type={}\n{}'.format(stype, status)
        self.locationtext.text = C4 + self.gps_status
        
    @mainthread
    def on_pause(self):
        gps.stop()
        return True

    def on_resume(self):
        gps.start(1000, 0)
        pass

class MyApp(App):
    def build(self):
        return InterfaceManager()


if __name__ in ('__main__', '__android__'):
    MyApp().run()
